import os
from openpyxl import load_workbook, Workbook # Importa load_workbook para abrir arquivos existentes
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter # Para trabalhar com letras de coluna


MAP_PORTAIS_PADRAO = {
    "meu_portal_antigo": "Meu Portal Oficial",
    "portal_da_web": "Portal Web Services",
    "portal_xpto": "Portal XPTO S.A.",
    "comprasnet gov": "Comprasnet Gov", # Ajustado para o nome completo padronizado
    "licitacoes_online": "Licitações Online",
    "bancodobrasil": "Banco do Brasil",
    "bll compras": "BLL Compras",
    "bnccompras": "BNC Compras",
    # ADICIONE AQUI MAIS MAPEAMENTOS CONFORME NECESSÁRIO PARA SEUS PORTAIS
}

# Mapeamento para padronizar nomes de clientes.
# Chave: nome do cliente como pode aparecer na planilha (em minúsculas para busca case-insensitive).
# Valor: nome padronizado que aparecerá no Excel.
MAP_CLIENTES_PADRAO = {
    "cliente_a_ltda": "Cliente A Ltda.",
    "cliente_b_sa": "Cliente B S.A.",
    "cliente_teste": "Cliente Teste ABC",
    "grupo_alpha": "Grupo Alpha Serviços",
    "prefeitura_sp": "Prefeitura de São Paulo",
    "construtora_x": "Construtora X Ltda",
    "licitec": "Licitiec",
    "3trend": "3Trend",
    "air liquide": "Air Liquide",
    "hexis": "Hexis",
    "cabala": "Cabala",
    "carlos andrade": "Carlos Andrade",
    "vimazi":"Vimazi Máquinas"
    # ADICIONE AQUI MAIS MAPEAMENTOS CONFORME NECESSÁRIO PARA SEUS CLIENTES
}

# Nome fixo do colaborador a ser adicionado em todas as linhas.
NOME_COLABORADOR = "João" # <-- LEMBRE-SE DE ALTERAR ESTE NOME PARA O COLABORADOR CORRETO

# --- Caminho da Pasta de Relatórios Existentes ---
PASTA_RELATORIOS_EXISTENTES = r"C:\Users\tec01\Desktop\contagem_py\retornos"

# --- Função para Padronizar um Único Arquivo Excel usando OpenPyXL ---
def padronizar_arquivo_excel_openpyxl(caminho_completo_arquivo: str):
    print(f"Processando: {os.path.basename(caminho_completo_arquivo)}")
    
    try:
        # Carrega o arquivo de trabalho
        wb = load_workbook(caminho_completo_arquivo)
        ws = wb.active 

        # --- Ler Dados Existentes e Determinar Colunas ---
        header_row = [cell.value for cell in ws[1]]
        
        # Mapeia cabeçalhos para seus índices de coluna (0-baseado para facilidade)
        col_indices = {header.lower() if header else "": i for i, header in enumerate(header_row)}
        
        # Lista para armazenar todos os dados das linhas (incluindo o cabeçalho)
        all_data = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx == 0: # Cabeçalho
                all_data.append(row) # Guarda o cabeçalho original, mas será substituído
            else: # Linhas de dados
                row_dict = {}
                for c_idx, cell_value in enumerate(row):
                    if c_idx < len(header_row) and header_row[c_idx] is not None:
                        row_dict[header_row[c_idx].lower()] = cell_value
                all_data.append(row_dict)

        # --- Limpar a Planilha para Reescrever com Nova Ordem e Dados ---
        # Deleta todas as linhas para reescrever
        ws.delete_rows(1, ws.max_row) 

        # --- Definir a Nova Ordem e Nomes dos Cabeçalhos ---
        # Coluna 'Colaborador' será a Coluna B (índice 1 no 0-based)
        new_headers = ['Data', 'Colaborador', 'Portal', 'Cliente', 'Quantidade de Processos Únicos'] # Nomes finais das colunas

        # --- Aplicar Formatação de Cabeçalho ---
        dark_blue_fill = PatternFill(start_color="00000080", end_color="00000080", fill_type="solid")
        white_bold_font = Font(color="FFFFFFFF", bold=True)
        
        ws.append(new_headers) # Adiciona o novo cabeçalho
        for cell in ws[1]: # Aplica estilo à primeira linha (novo cabeçalho)
            cell.fill = dark_blue_fill
            cell.font = white_bold_font

        # --- Preencher a Planilha com os Dados Reordenados e Padronizados ---
        for row_data in all_data[1:]: # Pula o cabeçalho original na leitura
            data_valor = row_data.get('data')
            portal_valor = row_data.get('portal')
            cliente_valor = row_data.get('cliente')
            quantidade_valor = row_data.get('quantidade de processos únicos') # Assumindo este nome agora

            # Aplica padronização
            portal_padronizado = MAP_PORTAIS_PADRAO.get(str(portal_valor).lower(), portal_valor) if portal_valor is not None else None
            cliente_padronizado = MAP_CLIENTES_PADRAO.get(str(cliente_valor).lower(), cliente_valor) if cliente_valor is not None else None
            
            # Adiciona a linha na nova ordem
            ws.append([
                data_valor,           # Coluna A
                NOME_COLABORADOR,     # Coluna B
                portal_padronizado,   # Coluna C
                cliente_padronizado,  # Coluna D
                quantidade_valor      # Coluna E
            ])
        
        # --- Salvar o arquivo padronizado ---
        wb.save(caminho_completo_arquivo) # Sobrescreve o arquivo original
        print(f"  '{os.path.basename(caminho_completo_arquivo)}' padronizado e salvo.")

    except FileNotFoundError:
        print(f"  Erro: Arquivo '{os.path.basename(caminho_completo_arquivo)}' não encontrado. Pulando.")
    except Exception as e:
        print(f"  Erro inesperado ao processar '{os.path.basename(caminho_completo_arquivo)}': {e}")

# --- Execução Principal do Script de Padronização ---
if __name__ == "__main__":
    print(f"Iniciando a padronização dos arquivos Excel na pasta: '{PASTA_RELATORIOS_EXISTENTES}'\n")
    
    # Verifica se a pasta existe
    if not os.path.exists(PASTA_RELATORIOS_EXISTENTES):
        print(f"Erro: A pasta '{PASTA_RELATORIOS_EXISTENTES}' não foi encontrada. Verifique o caminho.")
    else:
        # Lista todos os arquivos Excel (.xlsx) na pasta
        arquivos_excel = [f for f in os.listdir(PASTA_RELATORIOS_EXISTENTES) if f.endswith('.xlsx')] 
        
        if not arquivos_excel:
            print("Nenhum arquivo Excel .xlsx encontrado na pasta para padronizar.")
        else:
            for arquivo in arquivos_excel:
                caminho_completo = os.path.join(PASTA_RELATORIOS_EXISTENTES, arquivo)
                padronizar_arquivo_excel_openpyxl(caminho_completo)
        
    print("\nProcesso de padronização de arquivos concluído!")