import os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pdfplumber
import re
from collections import defaultdict
import json



CLIENTES_EDITAIS = [
    "Divena", "Danfoss", "IHO Soluções", "MCS", "ReawPlay",
    "Ortopedia Jaguaribe", "Educantes", "3Trend", "Inside",
    "Fioravant", "Benefício Certo", "TOTVS","TOTVs", "Unicontrols",
    "Vimazi","Covazi"
]

NOME_COLABORADOR = "João"

def carregar_json(caminho_arquivo):
    
    try:
        with open(caminho_arquivo, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Aviso: Arquivo '{caminho_arquivo}' não encontrado.")
        return {}
    except json.JSONDecodeError:
        print(f"Erro: Não foi possível decodificar o JSON do arquivo '{caminho_arquivo}'. Verifique a formatação.")
        return {}

def gerar_relatorio():
    pasta_raiz_busca = r"C:\Users\tec01\Desktop\Sites"

    MAP_CLIENTES_PADRAO = carregar_json('padrao_clientes.json')
    MAP_PORTAIS_PADRAO = carregar_json('padrao_portais.json')
    
    hoje = date.today()
    data_formatada = hoje.strftime("%Y-%m-%d")
    excel_saida = fr"C:\Users\tec01\Desktop\pasta_de_trabalho\contagem_pdf\retornos\contagem_{data_formatada}.xlsx"

    pasta_destino = os.path.dirname(excel_saida)
    os.makedirs(pasta_destino, exist_ok=True)

    dados_a_registrar = []
    arquivos_encontrados_hoje = False

    pregao_pattern = re.compile(r"pregão|pregao", re.IGNORECASE)
    data_numero_uasg_pattern = re.compile(r"\d{8}_\d{9}_\d{6}", re.IGNORECASE)
    
    id_extraction_patterns = [
        re.compile(r"\d+_\s*(\d+_\d+)", re.IGNORECASE), 
        re.compile(r"pregao\s*(\d+(?:_\d+)*)", re.IGNORECASE), 
        re.compile(r"edital\s*(\d{4}[-/_]?\d+)", re.IGNORECASE),
        re.compile(r"licitacao(?:-|\s)(\d+)", re.IGNORECASE),
        re.compile(r".*?_(\d+)_\d+\.pdf", re.IGNORECASE),
        re.compile(r"(\d{8}_\d{9}_\d{6})", re.IGNORECASE)
    ]

    for raiz, _, arquivos in os.walk(pasta_raiz_busca):
        for arquivo in arquivos:
            if arquivo.lower().endswith(".pdf"):
                caminho_completo = os.path.join(raiz, arquivo)
                try:
                    timestamp = os.path.getmtime(caminho_completo)
                    data_mod = datetime.fromtimestamp(timestamp)

                    if data_mod.date() == hoje:
                        arquivos_encontrados_hoje = True
                        
                        arquivo_lower = arquivo.lower()
                        
                        is_pregao_no_titulo = bool(pregao_pattern.search(arquivo_lower))
                        is_data_numero_uasg_pattern = bool(data_numero_uasg_pattern.search(arquivo_lower))

                        if is_pregao_no_titulo or is_data_numero_uasg_pattern:
                            relative_path = os.path.relpath(raiz, pasta_raiz_busca)
                            partes_caminho = os.path.normpath(relative_path).split(os.sep)
                            
                            diretorio_portal = "Portal Desconhecido"
                            diretorio_cliente = "Cliente Desconhecido"
                            
                            if len(partes_caminho) >= 1:
                                diretorio_portal = partes_caminho[0]
                            if len(partes_caminho) >= 2:
                                diretorio_cliente = partes_caminho[1]

                            if len(partes_caminho) == 1 and diretorio_portal.lower() not in MAP_PORTAIS_PADRAO:
                                diretorio_cliente = diretorio_portal
                                diretorio_portal = "Portal Desconhecido"
                                
                            if "licitar digital" in diretorio_portal.lower():
                                diretorio_portal = "Licitar Digital"
                                if len(partes_caminho) >= 2:
                                    diretorio_cliente = partes_caminho[1]
                            
                            portal_padronizado = MAP_PORTAIS_PADRAO.get(diretorio_portal.lower(), diretorio_portal)
                            cliente_padronizado = MAP_CLIENTES_PADRAO.get(diretorio_cliente.lower(), diretorio_cliente)

                            tipo_de_contrato = "Regular"
                            if cliente_padronizado in CLIENTES_EDITAIS:
                                tipo_de_contrato = "Editais"
                            
                            identificador_processo = None
                            for pattern in id_extraction_patterns:
                                match_id = pattern.search(arquivo_lower)
                                if match_id:
                                    identificador_processo = match_id.group(1)
                                    break
                            
                            if identificador_processo:
                                dados_a_registrar.append({
                                    'data': data_formatada,
                                    'colaborador': NOME_COLABORADOR,
                                    'portal': portal_padronizado,
                                    'cliente': cliente_padronizado,
                                    'tipo_contrato': tipo_de_contrato,
                                    'id': identificador_processo
                                })

                except Exception as e:
                    print(f"Erro ao processar '{caminho_completo}': {e}")
    
    if dados_a_registrar:
        contagem_consolidada = defaultdict(int)
        for item in dados_a_registrar:
            chave = (item['data'], item['colaborador'], item['portal'], item['cliente'], item['tipo_contrato'], item['id'])
            contagem_consolidada[chave] += 1

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Licitações"

        dark_blue_fill = PatternFill(start_color="00000080", end_color="00000080", fill_type="solid")
        white_bold_font = Font(color="FFFFFFFF", bold=True)

        headers = ['Data', 'Colaborador', 'Portal', 'Cliente', 'Tipo de Contrato', 'ID', 'Quantidade de Licitações']
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = dark_blue_fill
            cell.font = white_bold_font

        for chave, quantidade in contagem_consolidada.items():
            data, colaborador, portal, cliente, tipo_contrato, identificador = chave
            ws.append([data, colaborador, portal, cliente, tipo_contrato, identificador, quantidade])
            
        wb.save(excel_saida)
        print(f"\n✅ Relatório gerado em: {excel_saida}")
    else:
        if arquivos_encontrados_hoje:
            print("⚠️ Nenhum arquivo 'carinha' que se encaixe nos critérios e que possua ID foi localizado.")
        else:
            print(f"ℹ️ Nenhum PDF modificado hoje foi encontrado no diretório de busca.")


if __name__ == "__main__":
    print(f"Iniciando a geração do relatório diário para a data de hoje: {date.today().strftime('%Y-%m-%d')}...\n")
    gerar_relatorio()
    print("\n✅ Processo de geração de relatório diário concluído.")